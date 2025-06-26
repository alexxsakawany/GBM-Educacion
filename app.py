# bibliotecas
from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = 'data.xlsx'

# Verificação da existência do ficheiro excel, senão criar.

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Candidaturas'
    ws.append(['Nome', 'Email', 'Telefone', 'Matricula', 'Classe', 'Curso'])
    wb.save(EXCEL_FILE)


@app.route('/')
def home():
    return render_template('main.html')


@app.route('/inscricao')
def form():
    return render_template('inscricao.html', mostrar_popup=False)


@app.route('/sobre')
def about():
    return render_template('sobre.html')


@app.route('/contactos')
def contactos():
    return render_template('contactos.html')


@app.route('/send', methods=['POST'])
def send():
    nome = request.form['nome']
    email = request.form['email']
    telefone = request.form['telefone']
    matricula = request.form['matricula']
    classe = request.form['classe']
    curso = request.form['curso']

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nome, email, telefone, matricula, classe, curso])

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return "<h3>Erro: o arquivo está em uso ou protegido. Fecha o Excel e tenta de novo.</h3>"

    wb.save(EXCEL_FILE)

    
    return render_template('inscricao.html', mostrar_popup=True, nome=nome)


if __name__ == '__main__':
    app.run(debug=True)
